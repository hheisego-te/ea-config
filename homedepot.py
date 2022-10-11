import time
import openpyxl
from openpyxl.styles import Font, Alignment
from progress.bar import Bar
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, WebDriverException
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# Constants & Variables
C_PWD = 'welcome'

# Date & Time
def timestamp():
    date_time_now = datetime.now()
    #dt_string = date_time_now.strftime("%m/%d/%H:%M:%S")
    return date_time_now.strftime("%m/%d/%H:%M:%S")

def dump_logs(d_logs):
    d_logs = str(d_logs)
    file_name = 'dump' + timestamp().replace('/', '_')[:8] + '.log'
    f = open(file_name, 'a+')  # open file in append mode
    f.write(d_logs)
    f.close()

# Selenium
options = Options()
#options.add_argument('--headless')
driver = webdriver.Firefox(options=options) #executable_path="C:\\Users\\Helmut\\Desktop\\geckodriver.exe",
#wait = WebDriverWait(driver, 7)
action = ActionChains(driver)

# Openpyxl
config_file = openpyxl.load_workbook('config.xlsx')
data_sheet = config_file["Config"]
#logs_tab = "Logs("+ dt_string +")"
#logs_sheet = config_file.create_sheet(logs_tab)
#logs_sheet = config_file[logs_tab]

# Progress Bar
bar = Bar('Configuring ', max=(data_sheet.max_row - 1))

def login(host_ip, password):

    portal = 'https://' + host_ip
    dump_logs(d_logs=portal)

    try:
        driver.set_page_load_timeout(11)
        driver.get(portal)
        time.sleep(1.77)
        driver.find_element(By.NAME, 'username').send_keys('admin')
        time.sleep(0.77)
        driver.find_element(By.NAME, "password").send_keys(password)
        time.sleep(0.77)
        driver.find_element(By.CSS_SELECTOR, '.btn-outline-default').submit()
        time.sleep(2.77)

        return True

    except TimeoutException as ex:

        print(" Host no reachable: " + portal)
        dump_logs(d_logs=ex)

        return False

    except(NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, WebDriverException) as ex:
        pass

        step = "\nNetwork Conf " + portal
        dump_logs(d_logs=step)
        dump_logs(d_logs=ex)
        return True


def initial_setup(host_ip, new_password, accgroup_token):

    status = ''
    logged = login(host_ip=host_ip, password='welcome')

    if logged is True:

        status += "\n" + timestamp() + "-> Enterprise Agent Reachable "

        try:
            time.sleep(1.77)
            driver.find_element(By.NAME, "originalPassword").send_keys(C_PWD)
            time.sleep(0.77)
            driver.find_element(By.NAME, "newPassword").send_keys(new_password)
            time.sleep(0.77)
            driver.find_element(By.NAME, "confirmPassword").send_keys(new_password)
            time.sleep(3.33)
            #change_password = WebDriverWait(driver, 17).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn:nth-child(5)")))
            #change_password.click()
            driver.find_element(By.CSS_SELECTOR, "button.btn:nth-child(5)").submit()
            time.sleep(1.77)

            status += "\n" + timestamp() + "-> Original Password Has Changed Successfully "

        except(NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException) as ex:
            pass
            #print(ex)
            dump_logs(d_logs=ex)

            status += "\n" + timestamp() + "-> Could not Change Original Password"

        try:
            time.sleep(1.77)
            driver.find_element(By.NAME, "accountToken").send_keys(accgroup_token)
            time.sleep(1.77)
            next_button = WebDriverWait(driver, 7).until(EC.element_to_be_clickable((By.ID, "setupButtonNext")))
            next_button.click()
            time.sleep(2.77)##setupButtonNext


            status += "\n" + timestamp() + "-> Account Group Token Changed Successfully "

        except(NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException) as ex:
            pass
            #print(ex)
            dump_logs(d_logs=ex)

        return status

    else:

        status += "\n" + timestamp() + "-> Enterprise Agent Does not Reachable "

        return status


def network_setup(host_ip, hostname, new_password, proxy, proxy_port, proxy_list):

    status = ''

    try:

        time.sleep(0.77)
        driver.switch_to.new_window('tab')
        tabs = driver.window_handles
        time.sleep(0.77)
        driver.switch_to.window(tabs[0])
        time.sleep(0.77)
        driver.close()
        time.sleep(0.77)
        driver.switch_to.window(tabs[1])
        time.sleep(1.77)

        logged = login(host_ip=host_ip, password=new_password)

        if logged is True:

            time.sleep(0.77)
            driver.find_element(By.LINK_TEXT, "Network").click()
            time.sleep(1.77)
            driver.find_element(By.ID, "hostname").clear()
            time.sleep(0.77)
            driver.find_element(By.ID, "hostname").send_keys(hostname)
            time.sleep(0.77)

            # Proxy
            driver.find_element(By.ID, "proxy-type-label-static").click()
            time.sleep(0.77)
            driver.find_element(By.NAME, "proxy-host").send_keys(proxy)
            time.sleep(0.77)
            driver.find_element(By.NAME, "proxy-port").send_keys(proxy_port)
            time.sleep(0.77)
            driver.find_element(By.ID, "bypass-list-input").send_keys(proxy_list)
            time.sleep(0.77)
            driver.find_element(By.CLASS_NAME, "input-group-append").click()
            time.sleep(0.77)
            # Done Network
            driver.find_element(By.ID, "submit-form").submit()
            time.sleep(5.55)

            status += "\n" + timestamp() + "-> Network Setup Complete "

        else:

            status += "\n" + timestamp() + "-> Network Setup Fail "

        return status

    except(TimeoutException, NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException) as ex:

        status += "\n" + timestamp() + "-> Network Setup Fail"
        dump_logs(d_logs=ex)

        return False


start_time = time.perf_counter()

for ea in data_sheet.iter_rows(min_col=1, max_col=7, min_row=2):

    log_output = ''
    bar.next()

    first_part = initial_setup(host_ip=ea[0].value, new_password=ea[1].value, accgroup_token=ea[2].value)
    time.sleep(1.7)

    if "Enterprise Agent Reachable" in first_part:

        log_output += str(first_part)

        second_part = network_setup(host_ip=ea[0].value, new_password=ea[1].value, hostname=ea[3].value, proxy=ea[5].value, proxy_port=ea[6].value, proxy_list="google.com;cisco.com;thousandeyes.com")

        if second_part:

            log_output += str(second_part)

        else:

            log_output += str(second_part)

    else:

        log_output += first_part
        data_sheet.cell(row=ea[0].row, column=1).font = Font(color="00FF0000")

    data_sheet.cell(row=ea[0].row, column=11, value=None)
    data_sheet.cell(row=ea[0].row, column=11, value=log_output).alignment = Alignment(shrink_to_fit=False, wrapText=True, horizontal='general')
    data_sheet.cell(row=ea[0].row, column=11).font = Font(color="00008B")
     # Green color="00339966"  + Red color="00FF0000" position -> ea[6].row
    config_file.save('config.xlsx')
    print(" Elapsed Time ", time.perf_counter() - start_time)

    if ea[0].value is None:
        break

driver.quit()
config_file.save('config.xlsx')
bar.finish()
print(" Total Elapsed Time ", time.perf_counter() - start_time)

#pip install progress
#pip install openpyxl
